import time

import pytest
from openpyxl.reader.excel import load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


@pytest.mark.usefixtures("startup")
class Baseclass:
    print("This is Baseclass")

    def sleep(self,s):
        time.sleep(s)

    def element_present(self,elemnt):
        assert self.driver.find_element(*elemnt).is_displayed() == True

    def dotclick(self, element):
        self.driver.find_element(*element).click()

    def button_enable(self, butt):
        assert self.driver.find_element(*butt).is_enabled() == True

    def button_disable(self, butt):
        assert self.driver.find_element(*butt).is_enabled() == False

    def element_wait(self, welement, sec):
        wait = WebDriverWait(self.driver, sec)
        wait.until(EC.visibility_of_element_located(welement))

    def move_to_elment(self,element):
        act = ActionChains(self.driver)
        refele = self.driver.find_element(*element)
        act.move_to_element(refele).perform()

    def datafromexcel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]

        row_count = sh.max_row
        col_count = sh.max_column

        for r in range(2, row_count + 1):
            row = []
            for j in range(1, col_count + 1):
                row.append(sh.cell(row=r, column=j).value)
            datalist.append(row)

        return datalist